import pandas as pd
import openpyxl
import os

files = [
    r"c:\Users\Nodar\Batumi Hills\Batumi_Hills_Apartments.xlsx",
    r"c:\Users\Nodar\2026 antigraviti\ფართების_ახალი_განაწილება.xlsx",
    r"c:\Users\Nodar\2026 antigraviti\ფართების_განაწილების_აქტი.xlsx"
]

output_file = r"c:\Users\Nodar\2026 antigraviti\scratch\excel_sheets_inspect.txt"

with open(output_file, "w", encoding="utf-8") as out:
    for f in files:
        if os.path.exists(f):
            out.write(f"\n=================== {os.path.basename(f)} ===================\n")
            try:
                wb = openpyxl.load_workbook(f, data_only=True)
                out.write(f"Sheets: {wb.sheetnames}\n")
                for sheet in wb.sheetnames:
                    out.write(f"\n--- Sheet: {sheet} ---\n")
                    ws = wb[sheet]
                    # read first 40 rows and 10 columns
                    data = []
                    for row in list(ws.iter_rows(values_only=True))[:50]:
                        if any(row):
                            data.append([str(x)[:30] if x is not None else "" for x in row[:12]])
                    df = pd.DataFrame(data)
                    out.write(df.to_string(index=False))
                    out.write("\n")
            except Exception as e:
                out.write(f"Error: {e}\n")
        else:
            out.write(f"Path not found: {f}\n")

print("Done inspecting excel files!")
