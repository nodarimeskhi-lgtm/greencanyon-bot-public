"""
export_to_sheets.py
Reads Settlement_Optimized_Diversity_Plan.xlsx and exports:
  1. A clean inventory CSV for the Sales Portal sync
  2. A full Google Sheets-ready XLSX (all sheets, all formulas preserved)
"""
import sys, io, shutil, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC  = r'C:\Users\Nodar\2026 antigraviti\05-Projects\Tsalka\1-Management\1-3-Budget\Settlement_Optimized_Diversity_Plan.xlsx'
OUT_XLSX = r'C:\Users\Nodar\2026 antigraviti\05-Projects\Tsalka\6-Sales\6-4-Marketing\GreenCanyon_ForGoogleSheets.xlsx'
OUT_CSV  = r'C:\Users\Nodar\2026 antigraviti\05-Projects\Tsalka\6-Sales\6-4-Marketing\green_canyon_inventory.csv'

# ── Step 1: Copy xlsx with formulas intact ──────────────────────────────────
shutil.copy2(SRC, OUT_XLSX)
print(f'✅ Sheets-ready XLSX: {os.path.basename(OUT_XLSX)}')

# ── Step 2: Open with openpyxl and add Status + Notes + управление columns ──
wb = load_workbook(OUT_XLSX)

MGMT_COLS = ['status', 'notes', 'buyer_name', 'sale_date', 'agent']
MGMT_HDR  = ['სტატუსი', 'შენიშვნა', 'მყიდველი', 'გაყ. თარიღი', 'აგენტი']
STATUS_VALS = {'free': '🟢 ხელმისაწვდომი', 'reserved': '🟡 დაჯავშნილი', 'sold': '⚫ გაყიდული'}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'  Processing sheet: {sheet_name}')

    # Find header row (row with "ნაკვეთის კოდი")
    header_row = None
    id_col = None
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or '').strip() == 'ნაკვეთის კოდი':
                header_row = cell.row
                id_col = cell.column
                break
        if header_row:
            break

    if not header_row:
        print(f'  ⚠️  Header not found in {sheet_name}, skipping')
        continue

    # Find last used column and add management block after it
    max_col = ws.max_column
    start_mgmt_col = max_col + 2  # Leave a gap column

    # Gap column header
    ws.cell(row=header_row, column=max_col+1, value='│')

    # Write management column headers
    for i, (col_key, col_hdr) in enumerate(zip(MGMT_COLS, MGMT_HDR)):
        c = ws.cell(row=header_row, column=start_mgmt_col + i, value=col_hdr)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='1F5C1F')
        c.alignment = Alignment(horizontal='center')

    # Fill default status for each data row
    data_rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        id_val = ws.cell(row=row_idx, column=id_col).value
        if id_val and str(id_val).strip() and not str(id_val).startswith('Unnamed'):
            id_str = str(id_val).strip()
            if re.match(r'^L[A-D]-\d+|^AH-', id_str):
                data_rows.append(row_idx)
                # Status cell with dropdown hint
                status_cell = ws.cell(row=row_idx, column=start_mgmt_col, value='free')
                status_cell.font = Font(color='1F5C1F', bold=True)
                # Notes
                ws.cell(row=row_idx, column=start_mgmt_col+1, value='')
                ws.cell(row=row_idx, column=start_mgmt_col+2, value='')
                ws.cell(row=row_idx, column=start_mgmt_col+3, value='')
                ws.cell(row=row_idx, column=start_mgmt_col+4, value='')

    # Column widths for management block
    ws.column_dimensions[get_column_letter(start_mgmt_col)].width = 14
    ws.column_dimensions[get_column_letter(start_mgmt_col+1)].width = 25
    ws.column_dimensions[get_column_letter(start_mgmt_col+2)].width = 18
    ws.column_dimensions[get_column_letter(start_mgmt_col+3)].width = 14
    ws.column_dimensions[get_column_letter(start_mgmt_col+4)].width = 16

    print(f'    Added management columns for {len(data_rows)} plots')

# ── Add Instructions sheet ────────────────────────────────────────────────────
if 'README' not in wb.sheetnames:
    rm = wb.create_sheet('README', 0)
    instructions = [
        ['GREEN CANYON — Sales Portal Management Sheet'],
        [''],
        ['📌 ინსტრუქცია: Sales Portal-ის განახლება'],
        [''],
        ['1. ეს ფაილი Google Sheets-ში ატვირთეთ (File → Import → Upload)'],
        ['2. File → Share → Publish to web → CSV format'],
        ['3. URL კოპირება'],
        ['4. Sales Portal → ⚙ Google Sheets კავშირი → URL ჩასმა'],
        [''],
        ['📊 სტატუსის values (status სვეტი):'],
        ['', 'free', '→ 🟢 ხელმისაწვდომი'],
        ['', 'reserved', '→ 🟡 დაჯავშნილი'],
        ['', 'sold', '→ ⚫ გაყიდული'],
        [''],
        ['⚠️  id სვეტი არ შეცვალოთ!'],
        [''],
        ['🔄 Portal განახლდება ყოველ 5 წუთში ავტომატურად'],
        ['   ან: Portal → 🔄 ახლავე განახლება ღილაკი'],
        [''],
        ['📞 Sales Portal URL:', 'https://green-canyon-sales-portal.netlify.app'],
    ]
    for row_data in instructions:
        rm.append(row_data)
    rm['A1'].font = Font(bold=True, size=14, color='1F5C1F')
    rm['A3'].font = Font(bold=True, size=12)
    rm.column_dimensions['A'].width = 50
    rm.column_dimensions['B'].width = 20
    rm.column_dimensions['C'].width = 30
    rm.sheet_properties.tabColor = '1F5C1F'

wb.save(OUT_XLSX)
print(f'✅ Management columns + README added')

# ── Step 3: Generate Portal sync CSV ─────────────────────────────────────────
print(f'\n📤 Generating Portal sync CSV...')
all_rows = []
xl = pd.ExcelFile(SRC)
zone_map = {
    'ახალი კლდის პირი': {'sheet_zones': ['LA', 'LC']},
    'ახალი ვაკის უბანი': {'sheet_zones': ['LB', 'LD']},
}

for sheet in xl.sheet_names:
    df = pd.read_excel(xl, sheet_name=sheet, header=1)
    # ID column detection
    id_col = None
    for col in df.columns:
        sample = df[col].dropna().astype(str).head(5)
        if any(re.match(r'^L[A-D]-\d+', v) for v in sample):
            id_col = col
            break
    if id_col is None:
        continue

    daily_col = None
    for col in df.columns:
        if 'დღიური' in str(col) or 'daily' in str(col).lower():
            daily_col = col
            break

    land_price_col = None
    for col in df.columns:
        if '1მ2' in str(col) or '1 მ2' in str(col):
            land_price_col = col
            break

    for _, row in df.iterrows():
        pid = str(row.get(id_col, '')).strip()
        if not pid or not re.match(r'^L[A-D]-\d+|^AH-', pid):
            continue
        zone = pid.split('-')[0]
        all_rows.append({
            'id': pid,
            'zone': zone,
            'status': 'free',
            'notes': '',
            'daily_rent': row.get(daily_col, '') if daily_col else '',
            'land_price': row.get(land_price_col, '') if land_price_col else '',
        })

# Apart hotels are now directly loaded from Excel

csv_df = pd.DataFrame(all_rows)
csv_df.to_csv(OUT_CSV, index=False, encoding='utf-8-sig')
print(f'✅ Portal CSV: {os.path.basename(OUT_CSV)} ({len(csv_df)} rows)')

print(f"""
╔══════════════════════════════════════════════════════╗
║  შედეგები მზადაა:                                    ║
║                                                      ║
║  1. Google Sheets XLSX:                              ║
║     GreenCanyon_ForGoogleSheets.xlsx                 ║
║     → ეს ატვირთეთ Google Sheets-ში                  ║
║     → ყველა ფორმულა + ახალი სტატუს სვეტები          ║
║                                                      ║
║  2. Portal Sync CSV:                                 ║
║     green_canyon_inventory.csv                       ║
║     → ეს Google Sheets-ში Import → Publish → CSV     ║
║     → URI Portal-ში ჩასვი                           ║
╚══════════════════════════════════════════════════════╝
""")
